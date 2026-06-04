import streamlit as st
import pandas as pd
from openpyxl import Workbook
from fpdf import FPDF
from io import BytesIO
import math

# Configuración de la App
st.set_page_config(page_title="Control de Pedidos Mhttps://github.com/cpmiguelandrada-collab/MaxPallet/edit/main/Cargapallet.pyadera", layout="wide")
st.image("MaxPallet.png", width=350)

# Estética: Verde Pallet y Marrón Madera
st.markdown("""
    <style>
    .main { background-color: #fdf5e6; }
    .stMetric { background-color: #ffffff; padding: 15px; border-radius: 10px; border-left: 5px solid #92D050; box-shadow: 2px 2px 5px rgba(0,0,0,0.1); }
    .stButton>button { background-color: #92D050; color: white; font-weight: bold; height: 3em; border: none; }
    h1, h2 { color: #3b2f2f; }
    </style>
    """, unsafe_allow_html=True)

# BASE DE DATOS ACTUALIZADA (Arlog corregido y distinción Tablas/Listones)
pallets_db = {
    "Modelo 2": {"t": [(1200, 70, 15, 6), (800, 70, 15, 5)], "l": [(70, 70, 95, 9)], "cl": 46},
    "Modelo 7": {"t": [(1000, 70, 15, 5), (1200, 70, 15, 6)], "l": [(70, 70, 95, 9)], "cl": 42},
    "Modelo 6": {"t": [(1200, 90, 17, 7), (1200, 120, 17, 3), (1000, 120, 17, 3)], "l": [(1000, 120, 75, 9)], "cl": 60},
    "Modelo 12": {"t": [(1500, 70, 15, 5), (1000, 70, 15, 7), (860, 70, 15, 5)], "l": [(70, 70, 90, 5), (110, 70, 90, 10)], "cl": 69},
    "Pampa": {"t": [(1200, 70, 15, 4), (1200, 70, 20, 6), (1000, 95, 20, 3)], "l": [(102, 95, 101, 9)], "cl": 60},
    "Arlog": {"t": [(1200, 120, 22, 6), (1200, 70, 22, 4), (1000, 100, 22, 3)], "l": [(70, 120, 120, 9)], "cl": 144},
    "Modelo 8": {"t": [(1500, 120, 15, 15)], "l": [(70, 120, 90, 12)], "cl": 54},
}

def generar_pdf_pro(df_t, df_l, pies_totales):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", 'B', 16)
    pdf.cell(200, 10, "PEDIDO DE MADERA - ORDEN DE PRODUCCION", ln=True, align='C')
    pdf.ln(10)
    
    pdf.set_font("Arial", 'B', 12)
    pdf.cell(100, 10, f"TOTAL PIES CUADRADOS: {pies_totales:.2f} ft")
    pdf.ln(10)

    # Sección Tablas
    pdf.set_fill_color(146, 208, 80)
    pdf.cell(190, 10, "TABLAS (CORTADAS A MEDIDA)", 1, 1, 'L', True)
    pdf.set_font("Arial", 'B', 10)
    pdf.cell(90, 8, "Medida (LxAxE)", 1)
    pdf.cell(50, 8, "Cantidad", 1)
    pdf.cell(50, 8, "Pies Cuad.", 1, 1)
    pdf.set_font("Arial", '', 10)
    for _, r in df_t.iterrows():
        pdf.cell(90, 8, str(r['Medida (LxAxE)']), 1)
        pdf.cell(50, 8, str(r['Cantidad']), 1)
        pdf.cell(50, 8, str(r['Pies Cuadrados']), 1, 1)
    
    pdf.ln(10)
    # Sección Listones
    pdf.set_fill_color(210, 180, 140)
    pdf.cell(190, 10, "LISTONES (TIRAS DE 3 METROS)", 1, 1, 'L', True)
    pdf.set_font("Arial", 'B', 10)
    pdf.cell(90, 8, "Escuadria (Ancho x Espesor)", 1)
    pdf.cell(50, 8, "Metros Lineales", 1)
    pdf.cell(50, 8, "Tiras 3m", 1, 1)
    pdf.set_font("Arial", '', 10)
    for _, r in df_l.iterrows():
        pdf.cell(90, 8, str(r['Escuadria (AxE)']), 1)
        pdf.cell(50, 8, str(r['Metros Totales']), 1)
        pdf.cell(50, 8, str(r['Tiras de 3m']), 1, 1)
    pdf_bytes = bytes(pdf.output(dest='S'))
    return pdf_bytes

st.title("🪵 Optimizador de Producción Pallets")

col_izq, col_der = st.columns([1, 2])
pedido = {}

with col_izq:
    st.subheader("🛒 Cargar Pedido")
    for nombre in pallets_db.keys():
        pedido[nombre] = st.number_input(f"Cant. {nombre}", min_value=0, step=1)

if any(pedido.values()):
    tablas_data = {}
    listones_raw = {}
    total_ft2 = 0

    for mod, cant in pedido.items():
        if cant > 0:
            # Procesar Tablas
            for l, a, e, c in pallets_db[mod]["t"]:
                m = (l, a, e)
                v_pies = (l * a * e * c * cant) / 2354700
                tablas_data[m] = tablas_data.get(m, 0) + (c * cant)
                total_ft2 += v_pies
            # Procesar Listones
            for l, a, e, c in pallets_db[mod]["l"]:
                m = (a, e)
                listones_raw[m] = listones_raw.get(m, 0) + (l * c * cant)
                total_ft2 += (l * a * e * c * cant) / 2354700

    # Crear DataFrames para pantalla
    df_tablas = pd.DataFrame([
        {"Medida (LxAxE)": f"{k[0]}x{k[1]}x{k[2]}", "Cantidad": v, "Pies Cuadrados": round((k[0]*k[1]*k[2]*v)/2354700, 2)} 
        for k, v in tablas_data.items()
    ])
    
    listones_final = []
    for esc, mm in listones_raw.items():
        listones_final.append({
            "Escuadria (AxE)": f"{esc[0]} x {esc[1]}",
            "Metros Totales": round(mm/1000, 2),
            "Tiras de 3m": math.ceil(mm/3000)
        })
    df_listones = pd.DataFrame(listones_final)

    with col_der:
        st.subheader("📊 Totales Requeridos")
        st.metric("PIE CUADRADOS TOTALES (ft²)", f"{total_ft2:.2f}")
        
        st.write("**📝 Tablas a cortar (A medida)**")
        st.table(df_tablas)
        
        st.write("**🌲 Listones (Tiras de 3 metros)**")
        st.table(df_listones)

        st.divider()
        c1, c2 = st.columns(2)
        
        # Generar Excel
        buffer_ex = BytesIO()
        with pd.ExcelWriter(buffer_ex, engine='openpyxl') as writer:
            df_tablas.to_excel(writer, sheet_name='Tablas', index=False)
            df_listones.to_excel(writer, sheet_name='Listones_3m', index=False)
        c1.download_button("📂 Descargar y Abrir Excel", buffer_ex.getvalue(), "pedido_madera.xlsx")
        
        # Generar PDF
        pdf_b = generar_pdf_pro(df_tablas, df_listones, total_ft2)
        c2.download_button("📄 Generar PDF para Pedido", pdf_b, "pedido_aserradero.pdf")
else:
    with col_der:
        st.info("Ingresa cantidades para ver el resumen de madera y pies cuadrados.")
